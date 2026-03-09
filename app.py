from flask import Flask, jsonify, request, Response
from werkzeug.exceptions import RequestEntityTooLarge
import json
import base64
import io
import traceback
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import sys
import pandas as pd

# utils 경로 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils.vil_processor import process_vil_data, parse_target_current_from_filename
from utils.osc_processor import process_osc_data, parse_frequency_duty, get_preview_data
from utils.master_processor import process_master
from utils.trel_analysis import (
    analyze_single_file,
    get_preview_data as get_trel_preview,
    parse_vil_processed_for_voltage,
)

app = Flask(__name__)
CORS(app)


def csv_text_to_xlsx_bytes(csv_text: str) -> bytes:
    """CSV 문자열을 단일 시트 XLSX 바이너리로 변환 (Pandas Optimized)."""
    try:
        # comment='#' 옵션 추가: 메타데이터 라인 무시하고 데이터만 로드
        df = pd.read_csv(io.StringIO(csv_text), comment='#')
        output = io.BytesIO()
        # index=False로 인덱스 제외하고 저장
        df.to_excel(output, index=False, engine='openpyxl')
        return output.getvalue()
    except Exception as e:
        print(f"Excel Conversion Error: {e}", flush=True)
        # Fallback to manual conversion if pandas fails
        import csv
        import openpyxl
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        # Fallback: 메타데이터 라인('#')은 건너뛰고 데이터만 처리 시도
        lines = csv_text.splitlines()
        data_lines = [line for line in lines if not line.startswith('#')]
        
        reader = csv.reader(data_lines)
        for row in reader:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def parse_vil_uploaded_for_voltage(file_storage):
    """업로드된 VIL_processed 파일(CSV/XLSX)에서 (time_min, voltage) 목록 추출."""
    filename = (file_storage.filename or '').lower()
    raw = file_storage.read()
    
    if filename.endswith('.xlsx'):
        try:
            df = pd.read_excel(io.BytesIO(raw))
            # 컬럼명 확인 및 데이터 추출
            # 필요한 컬럼: Time (min), Voltage (V)
            # 대소문자 무시하고 찾기 위해 컬럼명 정규화
            df.columns = df.columns.str.strip()
            
            time_col = next((c for c in df.columns if 'Time (min)' in c), None)
            volt_col = next((c for c in df.columns if 'Voltage' in c), None)
            
            if time_col and volt_col:
                # (t, v) 튜플 리스트 반환
                return list(zip(df[time_col], df[volt_col]))
                
            # 컬럼 못 찾으면 첫 번째, 두 번째 컬럼 사용 (Fallback)
            if len(df.columns) >= 2:
                # 첫 행이 헤더일 수 있으니 숫자가 아닌 경우 제외
                df = df.apply(pd.to_numeric, errors='coerce').dropna()
                return list(zip(df.iloc[:, 0], df.iloc[:, 1]))
                
        except Exception:
            pass
        return []

    content = raw.decode('utf-8', errors='replace')
    _, data = parse_vil_processed_for_voltage(content)
    return data

# 요청 크기 제한
# MAX_UPLOAD_MB <= 0 이면 Flask 업로드 크기 제한을 비활성화한다(무제한).
MAX_UPLOAD_MB = int(os.environ.get('MAX_UPLOAD_MB', '0'))
if MAX_UPLOAD_MB > 0:
    app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

# Werkzeug 폼 파서 제한 해제
# - MAX_FORM_MEMORY_SIZE: multipart의 non-file 필드 메모리 제한
# - MAX_FORM_PARTS: multipart 파트(파일 개수 포함) 제한
app.config['MAX_FORM_MEMORY_SIZE'] = None
app.config['MAX_FORM_PARTS'] = None


# 413 에러 핸들러
@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(e):
    limit_text = f'{MAX_UPLOAD_MB}MB' if MAX_UPLOAD_MB > 0 else '무제한 설정(상위 프록시/서버 제한 가능)'
    return jsonify({
        'success': False,
        'error': f'요청 크기가 너무 큽니다. 전송하려는 파일이 너무 많거나 크기 때문일 수 있습니다. (현재 제한: {limit_text})'
    }), 413


@app.route('/')
def health_root():
    return jsonify({'status': 'ok', 'message': 'Flask backend is running'})


@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'message': 'Flask backend is running'})


@app.route('/api/process-vil', methods=['POST'])
def process_vil():
    """
    VIL 파일 처리 API
    - files: VIL이 포함된 CSV 파일들 (multipart/form-data)
    - 또는 file: 단일 파일
    """
    try:
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        files = [f for f in files if f and f.filename and not f.filename.startswith('._')]
        paths = request.form.getlist('paths') or []

        if not files:
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

        results = []
        processed_paths = set()
        for i, f in enumerate(files):
            filename = secure_filename(f.filename) or f.filename
            rel_path = paths[i] if i < len(paths) else filename

            if 'VIL' not in filename.upper():
                continue
            if rel_path in processed_paths:
                continue
            processed_paths.add(rel_path)

            target = parse_target_current_from_filename(filename)
            if target is None:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': '파일명에서 목표 전류를 추출할 수 없습니다.'
                })
                continue

            try:
                content = f.read().decode('utf-8', errors='replace')
                csv_out, time_shift, meta = process_vil_data(content, target, filename)
                xlsx_bytes = csv_text_to_xlsx_bytes(csv_out)
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'output_filename': meta.get('output_filename', filename.replace('.csv', '_processed.xlsx')),
                    'success': True,
                    'csv': csv_out,
                    'xlsx_b64': base64.b64encode(xlsx_bytes).decode('ascii'),
                    'time_shift_s': meta['time_shift_s'],
                    'time_shift_min': meta['time_shift_min'],
                    'target_current_ua': meta['target_current_ua'],
                    'original_points': meta['original_points'],
                    'filtered_points': meta['filtered_points']
                })
            except Exception as e:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        print(f"[process-vil] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/preview-osc', methods=['POST'])
def preview_osc():
    """
    오실로스코프 첫 파일 미리보기 (CH2 시각화용)
    - baseline 범위 지정을 위해 time_ns, CH2 데이터 반환
    """
    try:
        f = request.files.get('file')
        if not f or not f.filename or f.filename.startswith('._'):
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

        content = f.read().decode('utf-8', errors='replace')
        preview = get_preview_data(content)
        
        if 'error' in preview:
            return jsonify({'success': False, 'error': preview['error']}), 400
            
        freq, duty = parse_frequency_duty(f.filename or '')
        return jsonify({
            'success': True,
            'filename': f.filename,
            'time_ns': preview['time_ns'],
            'ch1': preview['ch1'],
            'ch2': preview['ch2'],
            'n_points': preview['n_points'],
            'frequency_hz': freq,
            'duty_fraction': duty,
        })
    except Exception as e:
        print(f"[preview-osc] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/process-osc', methods=['POST'])
def process_osc():
    """
    오실로스코프 TrEL 일괄 처리
    - files, paths, baseline_start_ns, baseline_end_ns
    """
    try:
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        files = [f for f in files if f and f.filename and not f.filename.startswith('._') and f.filename.lower().endswith('.csv')]
        paths = request.form.getlist('paths') or []

        baseline_start = float(request.form.get('baseline_start_ns', -500))
        baseline_end = float(request.form.get('baseline_end_ns', -100))
        norm_start = request.form.get('norm_start_ns', '').strip()
        norm_end = request.form.get('norm_end_ns', '').strip()
        norm_start_ns = float(norm_start) if norm_start else None
        norm_end_ns = float(norm_end) if norm_end else None

        if not files:
            return jsonify({'success': False, 'error': 'CSV 파일이 없습니다.'}), 400

        results = []
        processed_paths = set()
        for i, f in enumerate(files):
            filename = secure_filename(f.filename) or f.filename
            rel_path = paths[i] if i < len(paths) else filename
            if rel_path in processed_paths:
                continue
            processed_paths.add(rel_path)

            freq, duty = parse_frequency_duty(filename)
            if freq is None:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': '파일명에서 Frequency(Hz)를 추출할 수 없습니다.'
                })
                continue
            if duty is None:
                duty = 0.25  # 기본 25%

            try:
                content = f.read().decode('utf-8', errors='replace')
                csv_out, meta = process_osc_data(
                    content,
                    baseline_start_ns=baseline_start,
                    baseline_end_ns=baseline_end,
                    frequency_hz=freq,
                    duty_fraction=duty,
                    filename=filename,
                    norm_start_ns=norm_start_ns,
                    norm_end_ns=norm_end_ns,
                )
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'output_filename': meta['output_filename'],
                    'success': True,
                    'csv': csv_out,
                    'original_points': meta['original_points'],
                })
            except Exception as e:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        print(f"[process-osc] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/create-master', methods=['POST'])
def create_master():
    """
    VIL 기반 마스터 CSV 생성
    - vil_csv: VIL 처리된 CSV (Time min, Relative luminance)
    - vil_time_shift_min: VIL time shift (분)
    - files: *_TrEL.csv 파일들
    """
    try:
        print('[create-master] 요청 수신', flush=True)
        vil_csv = request.form.get('vil_csv', '').strip()
        vil_time_shift_min_str = request.form.get('vil_time_shift_min', '0').strip()
        vil_time_shift_min = float(vil_time_shift_min_str) if vil_time_shift_min_str else 0.0

        master_percents_str = request.form.get('master_percents', '100,90,80,70,60,50').strip()
        try:
            master_percents = [int(x.strip()) for x in master_percents_str.split(',') if x.strip()]
        except ValueError:
            master_percents = [100, 90, 80, 70, 60, 50]
        if not master_percents:
            master_percents = [100, 90, 80, 70, 60, 50]

        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        files = [f for f in files if f and f.filename and not f.filename.startswith('._') and '_TrEL.csv' in f.filename]

        if not vil_csv:
            return jsonify({'success': False, 'error': 'VIL 처리된 CSV가 필요합니다. 마스터 생성은 VIL 데이터를 기반으로 합니다.'}), 400
        if not files:
            return jsonify({'success': False, 'error': 'TrEL 처리된 CSV 파일이 없습니다. (_TrEL.csv)'}), 400

        files_data = []
        for f in files:
            content = f.read().decode('utf-8', errors='replace')
            files_data.append((f.filename, content))
        print(f'[create-master] TrEL {len(files_data)}개 로드, process_master 시작', flush=True)

        xlsx_bytes, _summary, metadata = process_master(vil_csv, vil_time_shift_min, files_data, percent_list=master_percents)
        print('[create-master] 완료', flush=True)

        # XLSX 파일 + X-Master-Metadata 헤더로 선택 파일 정보 전달
        resp = Response(
            xlsx_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=TrEL_Master.xlsx',
                'X-Master-Metadata': json.dumps(metadata, ensure_ascii=False),
                'Access-Control-Expose-Headers': 'X-Master-Metadata',
            }
        )
        return resp
    except RequestEntityTooLarge as e:
        return handle_request_entity_too_large(e)
    except ImportError as e:
        return jsonify({'success': False, 'error': f'openpyxl 설치 필요: pip install openpyxl'}), 500
    except Exception as e:
        print(f"[create-master] Error: {e}", flush=True)
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/trel-analysis-preview', methods=['POST'])
def trel_analysis_preview():
    """
    TrEL 배치 분석 미리보기 (첫 파일)
    - low_pct, high_pct, n_decay, decay_fit_start_us
    """
    try:
        f = request.files.get('file')
        if not f or not f.filename or f.filename.startswith('._'):
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400
        low_pct = float(request.form.get('low_pct', 0.1))
        high_pct = float(request.form.get('high_pct', 99))
        n_decay = int(request.form.get('n_decay', 2))
        decay_fit_start_us = float(request.form.get('decay_fit_start_us', 4.0))
        content = f.read().decode('utf-8', errors='replace')
        preview = get_trel_preview(content, low_pct, high_pct, n_decay, decay_fit_start_us)
        if preview.get('error'):
            return jsonify({'success': False, 'error': preview['error']}), 400
        return jsonify({'success': True, 'filename': f.filename, **preview})
    except Exception as e:
        print(f"[trel_analysis_preview] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/trel-analysis-batch', methods=['POST'])
def trel_analysis_batch():
    """
    TrEL 배치 분석 - Rise, Saturation, Decay 추출 후 Excel 저장 (Pandas Optimized)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl 설치 필요: pip install openpyxl'}), 500

    try:
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        vil_files = request.files.getlist('vil_files') or []
        files = [f for f in files if f and f.filename and not f.filename.startswith('._') and f.filename.lower().endswith('.csv') and ('_TrEL' in f.filename or 'TrEL' in f.filename) and 'VIL' not in f.filename.upper()]
        vil_files = [f for f in vil_files if f and f.filename and not f.filename.startswith('._') and 'VIL' in f.filename.upper() and '_processed' in f.filename]
        low_pct = float(request.form.get('low_pct', 0.1))
        high_pct = float(request.form.get('high_pct', 99))
        n_decay = int(request.form.get('n_decay', 2))
        decay_fit_start_us = float(request.form.get('decay_fit_start_us', 4.0))
        integration_limit_us = float(request.form.get('integration_limit_us', 5.0))
        baseline_start_us = float(request.form.get('baseline_start_us', 20.0))

        if not files:
            return jsonify({'success': False, 'error': 'CSV 파일이 없습니다.'}), 400

        vil_time_voltage = []
        for vf in vil_files:
            data = parse_vil_uploaded_for_voltage(vf)
            if data:
                vil_time_voltage.extend(data)
        if vil_time_voltage:
            seen_t = {}
            for t, v in sorted(vil_time_voltage, key=lambda x: x[0]):
                if t not in seen_t:
                    seen_t[t] = v
            vil_time_voltage = [(t, seen_t[t]) for t in sorted(seen_t.keys())]

        results = []
        preview_seed_popt = None
        for idx, f in enumerate(files):
            content = f.read().decode('utf-8', errors='replace')
            r = analyze_single_file(
                content, f.filename, low_pct, high_pct, n_decay,
                vil_time_voltage=vil_time_voltage if vil_time_voltage else None,
                decay_fit_start_us=decay_fit_start_us,
                decay_initial_params=preview_seed_popt if idx > 0 else None,
                integration_limit_us=integration_limit_us,
                baseline_start_us=baseline_start_us,
            )
            results.append(r)
            # First file (preview source) fit result is reused
            # as initial guess for remaining files.
            if preview_seed_popt is None and isinstance(r.get('popt'), list):
                preview_seed_popt = r.get('popt')

        has_voltage = any(row.get('voltage') is not None for row in results)
        has_capacitance = any(row.get('relative_capacitance') is not None for row in results)
        if has_voltage:
            results.sort(key=lambda r: (r.get('time_min') is None, r.get('time_min') or float('inf')))
        else:
            results.sort(key=lambda r: (r.get('after_duty') is None or r.get('after_duty') == '', r.get('after_duty') or ''))

        # Excel 생성 (Pandas)
        excel_rows = []
        first_col_name = 'Time (min)' if has_voltage else 'duty 뒤'
        
        for r in results:
            row = {}
            row[first_col_name] = r.get('time_min') if has_voltage else r.get('after_duty')
            row['fit_start_us (μs)'] = decay_fit_start_us
            
            if has_voltage:
                row['Voltage (V)'] = r.get('voltage')
            if has_capacitance:
                row['Rel. Capacitance (nC/cm²)'] = r.get('relative_capacitance')
                
            row['t_delay (μs)'] = r.get('t_delay')
            row['t_rise (μs)'] = r.get('t_rise')
            row['t_saturation (μs)'] = r.get('t_saturation')
            
            for i in range(1, n_decay + 1):
                row[f'tau_{i} (μs)'] = r.get(f'tau_{i}')
                row[f'f_{i}'] = r.get(f'f_{i}')
                
            row['tau_avg (μs)'] = r.get('tau_avg')
            excel_rows.append(row)
            
        df = pd.DataFrame(excel_rows)
        
        # 컬럼 순서 지정
        cols = [first_col_name, 'fit_start_us (μs)']
        if has_voltage: cols.append('Voltage (V)')
        if has_capacitance: cols.append('Rel. Capacitance (nC/cm²)')
        cols.extend(['t_delay (μs)', 't_rise (μs)', 't_saturation (μs)'])
        for i in range(1, n_decay + 1):
            cols.extend([f'tau_{i} (μs)', f'f_{i}'])
        cols.append('tau_avg (μs)')
        
        # 존재하는 컬럼만 선택
        cols = [c for c in cols if c in df.columns]
        df = df[cols]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='TrEL Analysis')
            ws = writer.sheets['TrEL Analysis']
            
            # 헤더 볼드 처리 및 컬럼 너비 조정
            for cell in ws[1]:
                cell.font = Font(bold=True)
                
            for i, col in enumerate(cols, 1):
                w = 18 if i == 1 and not has_voltage else 14
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=TrEL_Analysis.xlsx'}
        )
    except Exception as e:
        print(f"[trel_analysis_batch] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port_str = os.environ.get('PORT')
    if port_str:
        port = int(port_str)
    elif len(sys.argv) > 1:
        port = int(sys.argv[1])
    else:
        port = 8080

    host = os.environ.get('HOST', '0.0.0.0')

    print(f"Flask app starting...")
    print(f"Host: {host}, Port: {port}")
    sys.stdout.flush()

    try:
        print(f"Server started: http://{host}:{port}")
        sys.stdout.flush()
        app.run(debug=False, use_reloader=False, port=port, host=host)
    except OSError as e:
        error_msg = str(e)
        print(f"Port error: {error_msg}")
        sys.stdout.flush()
        if 'Address already in use' in error_msg or 'Port already in use' in error_msg:
            print(f"Port {port} is already in use.")
            print(f"Trying port 5001 instead...")
            sys.stdout.flush()
            app.run(debug=False, use_reloader=False, port=5001, host=host)
        else:
            raise
